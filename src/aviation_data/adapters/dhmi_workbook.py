from __future__ import annotations

from datetime import date, datetime, time
from io import BytesIO
from pathlib import Path
from typing import Any

PROFILE = "dhmi_workbook_v1"


def _json_value(value: Any) -> Any:
    if isinstance(value, (date, datetime, time)):
        return value.isoformat()
    return value


def _display(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return f"{value:.12g}"
    return str(value).replace("\n", " ").strip()


def _escape(value: Any) -> str:
    return _display(value).replace("\\", "\\\\").replace("|", "\\|")


def _merged_values(sheet: Any) -> dict[tuple[int, int], Any]:
    values: dict[tuple[int, int], Any] = {}
    for merged in sheet.merged_cells.ranges:
        value = sheet.cell(merged.min_row, merged.min_col).value
        for row in range(merged.min_row, merged.max_row + 1):
            for column in range(merged.min_col, merged.max_col + 1):
                values[(row, column)] = value
    return values


def _row_values(sheet: Any, row: int, width: int, merged: dict[tuple[int, int], Any]) -> list[Any]:
    return [
        merged.get((row, column), sheet.cell(row, column).value) for column in range(1, width + 1)
    ]


def _header_end(rows: list[list[Any]], title_index: int) -> int:
    for index in range(title_index + 1, len(rows)):
        numeric = sum(
            isinstance(value, (int, float)) and not isinstance(value, bool) for value in rows[index]
        )
        if numeric >= 1:
            return index
    raise ValueError("worksheet has no detectable data rows")


def _column_names(header_rows: list[list[Any]]) -> tuple[list[str], list[list[str]]]:
    filled = []
    for row in header_rows:
        current = ""
        output = []
        for value in row:
            text = _display(value)
            if text:
                current = text
            output.append(current)
        filled.append(output)
    hierarchy = []
    for column in range(len(filled[0])):
        values = []
        for row in filled:
            value = row[column]
            if value and (not values or value != values[-1]):
                values.append(value)
        hierarchy.append(values or [f"Column {column + 1}"])
    counts: dict[str, int] = {}
    names = []
    for values in hierarchy:
        base = " / ".join(values)
        counts[base] = counts.get(base, 0) + 1
        names.append(base if counts[base] == 1 else f"{base} ({counts[base]})")
    return names, hierarchy


def _row_type(values: list[Any], *, data_started: bool) -> str:
    first = _display(values[0]).casefold() if values else ""
    numeric = any(
        isinstance(value, (int, float)) and not isinstance(value, bool) for value in values[1:]
    )
    if "toplam" in first or "total" in first:
        return "total"
    if numeric:
        return "data"
    if data_started and any(_display(value) for value in values):
        return "note"
    return "empty"


def extract_dhmi_workbook(path: Path) -> tuple[str, dict[str, Any]]:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("dhmi_workbook_v1 extraction requires the 'formats' extra") from exc

    payload = path.read_bytes()
    formulas = openpyxl.load_workbook(BytesIO(payload), data_only=False, read_only=False)
    cached = openpyxl.load_workbook(BytesIO(payload), data_only=True, read_only=False)
    markdown: list[str] = []
    structured_sheets = []
    parquet_sheets: dict[str, list[list[str]]] = {}
    try:
        for formula_sheet in formulas.worksheets:
            cached_sheet = cached[formula_sheet.title]
            merged = _merged_values(formula_sheet)
            nonempty_columns = [
                column
                for column in range(1, formula_sheet.max_column + 1)
                if any(
                    _display(
                        merged.get(
                            (row, column),
                            formula_sheet.cell(row, column).value,
                        )
                    )
                    for row in range(1, formula_sheet.max_row + 1)
                )
            ]
            if not nonempty_columns:
                continue
            width = max(nonempty_columns)
            rows = [
                _row_values(formula_sheet, row, width, merged)
                for row in range(1, formula_sheet.max_row + 1)
            ]
            populated = [index for index, row in enumerate(rows) if any(_display(v) for v in row)]
            if not populated:
                continue
            title_index = populated[0]
            title = next(
                (_display(value) for value in rows[title_index] if _display(value)),
                formula_sheet.title,
            )
            header_end = _header_end(rows, title_index)
            header_rows = rows[title_index + 1 : header_end]
            if not header_rows:
                raise ValueError(f"worksheet {formula_sheet.title!r} has no header rows")
            names, hierarchy = _column_names(header_rows)
            output_rows = []
            table_rows = [names]
            notes = []
            data_started = False
            for row_index in range(header_end + 1, formula_sheet.max_row + 1):
                values = _row_values(formula_sheet, row_index, width, merged)
                kind = _row_type(values, data_started=data_started)
                if kind == "empty":
                    continue
                data_started = data_started or kind in {"data", "total"}
                if kind == "note":
                    note = " ".join(_display(value) for value in values if _display(value))
                    notes.append({"row_number": row_index, "text": note})
                    continue
                cells = {}
                display_row = []
                for column, name in enumerate(names, start=1):
                    formula_value = formula_sheet.cell(row_index, column).value
                    cached_value = cached_sheet.cell(row_index, column).value
                    value = cached_value if cached_value is not None else formula_value
                    cell = {"value": _json_value(value)}
                    if isinstance(formula_value, str) and formula_value.startswith("="):
                        cell["formula"] = formula_value
                        cell["cached_value"] = _json_value(cached_value)
                    cells[name] = cell
                    display_row.append(_display(value))
                output_rows.append(
                    {
                        "row_number": row_index,
                        "row_type": kind,
                        "cells": cells,
                    }
                )
                table_rows.append(display_row)
            if not output_rows:
                raise ValueError(f"worksheet {formula_sheet.title!r} is structurally empty")
            markdown.extend(
                [
                    f"# {title}",
                    "",
                    "| " + " | ".join(_escape(value) for value in names) + " |",
                    "| " + " | ".join(["---"] * len(names)) + " |",
                    *[
                        "| " + " | ".join(_escape(value) for value in row) + " |"
                        for row in table_rows[1:]
                    ],
                ]
            )
            if notes:
                markdown.extend(["", "## Notes", *[f"- {note['text']}" for note in notes]])
            markdown.append("")
            parquet_sheets[formula_sheet.title] = table_rows
            structured_sheets.append(
                {
                    "worksheet": formula_sheet.title,
                    "title": title,
                    "title_row": title_index + 1,
                    "header_rows": list(range(title_index + 2, header_end + 1)),
                    "columns": [
                        {
                            "column_index": index,
                            "name": name,
                            "header_hierarchy": values,
                        }
                        for index, (name, values) in enumerate(
                            zip(names, hierarchy, strict=True),
                            start=1,
                        )
                    ],
                    "rows": output_rows,
                    "notes": notes,
                }
            )
    finally:
        formulas.close()
        cached.close()
    if not structured_sheets:
        raise ValueError("workbook has no structurally usable worksheets")
    structured = {
        "schema_version": "1.0.0",
        "extraction_profile": PROFILE,
        "sheets": structured_sheets,
    }
    return "\n".join(markdown).strip(), {
        "extractor": "openpyxl",
        "extraction_profile": PROFILE,
        "sheets": parquet_sheets,
        "worksheet_count": len(structured_sheets),
        "_structured_artifact": {
            "filename": "dhmi_workbook.json",
            "key": "dhmi_workbook_json",
            "value": structured,
        },
    }
