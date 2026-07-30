from __future__ import annotations

import asyncio
import hashlib
from datetime import date
from pathlib import Path

import openpyxl
import pymupdf
import pytest
import yaml

from aviation_data.acquisition import detect_mime, fetch_sources
from aviation_data.adapters.dhmi_workbook import extract_dhmi_workbook
from aviation_data.adapters.easa_sections import extract_easa_section
from aviation_data.adapters.faa_sections import extract_faa_sections
from aviation_data.adapters.shgm_abbreviations import extract_shgm_abbreviations
from aviation_data.curation import _quota_report
from aviation_data.models import DocumentRecord, Language, RightsState, Topic
from aviation_data.registry import load_registry


def _local_registry(project: Path, pattern: str) -> Path:
    registry_path = project / "configs" / "sources.yaml"
    registry_path.parent.mkdir(parents=True, exist_ok=True)
    raw = {
        "project": {
            "name": "local-test",
            "contact": "test@example.test",
            "user_agent": "local-test/1",
        },
        "sources": [
            {
                "source_id": "local_test",
                "enabled": True,
                "adapter": "local_glob",
                "seed_urls": [pattern],
                "publisher": "Test",
                "source_family": "test",
                "authority_level": "fixture",
                "languages": ["en"],
                "topics": ["aircraft"],
                "expected_mime_types": ["application/pdf"],
                "native_format": "pdf",
                "update_cadence": "manual",
                "version_discovery": "checksum",
                "rights": {
                    "state": "open",
                    "license_id": "CC0-1.0",
                    "license_url": "https://example.test/license",
                    "terms_url": "https://example.test/terms",
                    "reviewed_on": "2026-07-30",
                    "attribution": "Test",
                    "release_source": True,
                    "release_derived_text": True,
                    "release_qa": True,
                },
            }
        ],
    }
    registry_path.write_text(yaml.safe_dump(raw), encoding="utf-8")
    return registry_path


def test_local_glob_is_sorted_optional_versioned_and_checksum_deduplicated(
    tmp_path: Path,
) -> None:
    project = tmp_path / "project"
    inputs = project / "datatoprocess"
    inputs.mkdir(parents=True)
    (inputs / "Test_b.pdf").write_bytes(b"%PDF-version-b")
    (inputs / "Test_a.pdf").write_bytes(b"%PDF-version-a")
    (inputs / "Test_a_renamed.pdf").write_bytes(b"%PDF-version-a")
    registry_path = _local_registry(project, "datatoprocess/Test*.pdf")
    registry = load_registry(registry_path)
    data_dir = project / "data"

    records, errors = asyncio.run(
        fetch_sources(
            registry,
            registry_path,
            data_dir,
            date(2026, 7, 30),
        )
    )

    assert not errors
    assert [record.canonical_url for record in records] == [
        "file:datatoprocess/Test_a.pdf",
        "file:datatoprocess/Test_b.pdf",
    ]
    assert len({record.sha256 for record in records}) == 2
    assert records[0].fetch_recipe["configured_glob"] == "datatoprocess/Test*.pdf"
    assert records[0].fetch_recipe["concrete_relative_path"] == ("datatoprocess/Test_a.pdf")
    assert records[0].fetch_recipe["checksum"] == records[0].sha256

    (inputs / "Test_a.pdf").write_bytes(b"%PDF-version-a-changed")
    records, errors = asyncio.run(
        fetch_sources(
            registry,
            registry_path,
            data_dir,
            date(2026, 7, 31),
        )
    )
    assert not errors
    assert len(records) == 3
    assert len({record.source_version for record in records}) == 3


def test_local_glob_no_match_is_nonfatal_and_traversal_is_rejected(tmp_path: Path) -> None:
    project = tmp_path / "project"
    registry_path = _local_registry(project, "datatoprocess/Missing*.pdf")
    registry = load_registry(registry_path)
    records, errors = asyncio.run(
        fetch_sources(registry, registry_path, project / "data", date(2026, 7, 30))
    )
    assert records == []
    assert errors == []

    unsafe_path = _local_registry(project, "datatoprocess/../*.pdf")
    unsafe = load_registry(unsafe_path)
    records, errors = asyncio.run(
        fetch_sources(unsafe, unsafe_path, project / "unsafe-data", date(2026, 7, 30))
    )
    assert records == []
    assert errors and "path traversal" in errors[0]["error"]


def test_xlsx_mime_uses_workbook_type_instead_of_zip() -> None:
    assert (
        detect_mime("sample.xlsx", b"PK\x03\x04workbook")
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


def _ratio_document(
    document_id: str,
    language: Language,
    tokens: int,
    *,
    releasable: bool,
) -> DocumentRecord:
    return DocumentRecord(
        document_id=document_id,
        document_version="v1",
        variant_group_id=f"variant-{document_id}",
        title=document_id,
        language=language,
        topics=[Topic.AIRCRAFT],
        publisher="Test",
        source_family=f"family-{document_id}",
        authority_level="fixture",
        source_record_id=f"source-{document_id}",
        source_url=f"file:{document_id}",
        native_mime="text/plain",
        native_format="text",
        license_id="test",
        attribution="Test",
        rights_state=RightsState.OPEN if releasable else RightsState.MANIFEST_ONLY,
        release_derived_text=releasable,
        release_qa=releasable,
        canonical_path=f"{document_id}.md",
        canonical_sha256="0" * 64,
        canonical_char_count=tokens,
        canonical_token_count=tokens,
    )


def test_language_ratios_are_observational_and_separate_qa_eligible_documents(
    tmp_path: Path,
) -> None:
    documents = [
        _ratio_document("open-en", Language.ENGLISH, 70, releasable=True),
        _ratio_document("open-tr", Language.TURKISH, 30, releasable=True),
        _ratio_document("restricted-tr", Language.TURKISH, 100, releasable=False),
    ]

    report = _quota_report(
        documents,
        {
            "language_token_reference": {"en": 0.7, "tr": 0.3},
            "language_observation_tolerance": 0.05,
            "minimum_topic_share": 0,
            "maximum_source_family_share": 1,
        },
        tmp_path,
    )

    observation = report["language_observation"]
    assert observation["blocking"] is False
    assert observation["views"]["all_accepted"]["token_shares"] == {
        "en": 0.35,
        "tr": 0.65,
    }
    assert observation["views"]["qa_eligible"]["token_shares"] == {
        "en": 0.7,
        "tr": 0.3,
    }
    assert observation["views"]["all_accepted"]["status"] == "outside_tolerance"
    assert observation["views"]["qa_eligible"]["status"] == "within_tolerance"
    assert not any(issue["code"] == "language_quota" for issue in report["quota_issues"])


def test_dhmi_profile_extracts_hierarchical_tables_formulas_totals_and_notes(
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "DHMI_fixture.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Traffic"
    sheet.append(["AIR TRAFFIC", None, None, None, None])
    sheet.append(["Airport", "2025", None, "2026", None])
    sheet.append([None, "Domestic", "Total", "Domestic", "Total"])
    sheet.append(["Alpha", 10, 10, 12, "=SUM(D4)"])
    sheet.append(["TOTAL", 10, 10, 12, 12])
    sheet.append(["Note: preliminary values", None, None, None, None])
    sheet.merge_cells("A1:E1")
    sheet.merge_cells("A2:A3")
    sheet.merge_cells("B2:C2")
    sheet.merge_cells("D2:E2")
    second = workbook.create_sheet("Cargo")
    second.append(["CARGO", None])
    second.append(["Airport", "Tonnes"])
    second.append(["Alpha", 4.5])
    workbook.save(workbook_path)

    markdown, layout = extract_dhmi_workbook(workbook_path)
    artifact = layout["_structured_artifact"]["value"]

    assert "# AIR TRAFFIC" in markdown
    assert "2025 / Domestic" in markdown
    assert "## Notes" in markdown
    assert artifact["schema_version"] == "1.0.0"
    assert len(artifact["sheets"]) == 2
    traffic = artifact["sheets"][0]
    assert [row["row_type"] for row in traffic["rows"]] == ["data", "total"]
    formula_cell = traffic["rows"][0]["cells"]["2026 / Total"]
    assert formula_cell["formula"] == "=SUM(D4)"
    assert "cached_value" in formula_cell
    assert traffic["notes"][0]["text"] == "Note: preliminary values"
    assert layout["sheets"]["Traffic"][0][1] == "2025 / Domestic"


def _save_pdf(path: Path, pages: list[list[tuple[tuple[float, float], str]]]) -> None:
    document = pymupdf.open()
    for rows in pages:
        page = document.new_page()
        for point, text in rows:
            page.insert_text(point, text)
    document.save(path)
    document.close()


def test_shgm_profile_parses_page_spanning_aliases_and_excludes_definitions(
    tmp_path: Path,
) -> None:
    path = tmp_path / "SHGM_fixture.pdf"
    _save_pdf(
        path,
        [
            [
                ((72, 72), "TANIMLAR VE KISALTMALAR"),
                ((72, 95), "MADDE 4"),
                ((72, 120), "a) ACC: Saha Kontrol Merkezi"),
            ],
            [
                ((72, 72), "(Area Control Center),"),
                ((72, 95), "b) HSKP (AUP): Hava Sahasi Kullanim Plani (Airspace Use Plan),"),
                ((72, 120), "c) General Definition: This is not an abbreviation,"),
                ((72, 145), "IKINCI BOLUM"),
                ((72, 170), "MADDE 5"),
            ],
        ],
    )

    markdown, layout = extract_shgm_abbreviations(path)
    artifact = layout["_structured_artifact"]["value"]

    assert artifact["mapping"]["ACC"] == "Saha Kontrol Merkezi"
    assert artifact["mapping"]["HSKP"] == "Hava Sahasi Kullanim Plani"
    assert artifact["mapping"]["AUP"] == "Hava Sahasi Kullanim Plani"
    assert artifact["entries"][0]["page_start"] == 1
    assert artifact["entries"][0]["page_end"] == 2
    assert "General Definition" not in markdown


def test_shgm_profile_fails_when_required_section_is_missing(tmp_path: Path) -> None:
    path = tmp_path / "SHGM_missing.pdf"
    _save_pdf(path, [[((72, 72), "MADDE 1"), ((72, 95), "No abbreviations here.")]])
    with pytest.raises(ValueError, match="heading not found"):
        extract_shgm_abbreviations(path)


def _section_pdf(path: Path) -> str:
    document = pymupdf.open()
    first = document.new_page()
    first.insert_text((72, 80), "SUBPART A")
    first.insert_text((72, 120), "CS-E 10 Applicability")
    first.insert_text((72, 150), "Parent section text.")
    first.insert_text((72, 230), "AMC E 10 Guidance")
    first.insert_text((72, 260), "Nested guidance text.")
    first.insert_text((72, 360), "CS-E 20 Terminology")
    first.insert_text((72, 390), "Next section text.")
    document.set_toc(
        [
            [1, "SUBPART A", 1, {"kind": 1, "page": 0, "to": pymupdf.Point(72, 80)}],
            [
                2,
                "CS-E 10 Applicability",
                1,
                {"kind": 1, "page": 0, "to": pymupdf.Point(72, 120)},
            ],
            [
                3,
                "AMC E 10 Guidance",
                1,
                {"kind": 1, "page": 0, "to": pymupdf.Point(72, 230)},
            ],
            [
                2,
                "CS-E 20 Terminology",
                1,
                {"kind": 1, "page": 0, "to": pymupdf.Point(72, 360)},
            ],
        ]
    )
    document.save(path)
    document.close()
    return hashlib.sha256(path.read_bytes()).hexdigest()


def test_easa_profile_is_deterministic_and_uses_coordinate_hierarchy(
    tmp_path: Path,
) -> None:
    path = tmp_path / "EASA_fixture.pdf"
    checksum = _section_pdf(path)
    seed_for_parent = next(
        seed
        for seed in range(100)
        if int(hashlib.sha256(f"{seed}:{checksum}".encode()).hexdigest(), 16) % 3 == 0
    )
    markdown, layout = extract_easa_section(
        path,
        seed=seed_for_parent,
        checksum=checksum,
    )
    repeated, _ = extract_easa_section(path, seed=seed_for_parent, checksum=checksum)
    artifact = layout["_structured_artifact"]["value"]

    assert markdown == repeated
    assert artifact["eligible_count"] == 3
    assert artifact["chosen_title"] == "CS-E 10 Applicability"
    assert "Parent section text" in markdown
    assert "Nested guidance text" in markdown
    assert "Next section text" not in markdown
    assert artifact["start_coordinate"]["y"] < artifact["end_coordinate"]["y"]

    other_seed = next(
        seed
        for seed in range(100)
        if int(hashlib.sha256(f"{seed}:{checksum}".encode()).hexdigest(), 16) % 3 != 0
    )
    _, other_layout = extract_easa_section(path, seed=other_seed, checksum=checksum)
    assert other_layout["_structured_artifact"]["value"]["chosen_title"] != artifact["chosen_title"]


def test_easa_profile_fails_without_eligible_toc(tmp_path: Path) -> None:
    path = tmp_path / "EASA_missing.pdf"
    _save_pdf(path, [[((72, 72), "No bookmarks")]])
    with pytest.raises(ValueError, match="bookmark TOC"):
        extract_easa_section(path, seed=1, checksum="0" * 64)


def test_faa_profile_avoids_contents_and_keeps_cross_page_continuation(
    tmp_path: Path,
) -> None:
    path = tmp_path / "FAA_fixture.pdf"
    document = pymupdf.open()
    contents = document.new_page()
    contents.insert_text((72, 72), "Contents")
    contents.insert_text((72, 100), "1 PURPOSE ........ 1")
    purpose = document.new_page()
    purpose.insert_text((72, 100), "1 PURPOSE.")
    purpose.insert_text((72, 130), "The purpose begins here.")
    purpose.insert_text((72, 300), "2 APPLICABILITY.")
    purpose.insert_text((72, 330), "The applicability section begins here.")
    continuation = document.new_page()
    continuation.insert_text((72, 30), "RUNNING HEADER")
    continuation.insert_text((72, 100), "Applicability continues on the next page.")
    continuation.insert_text((300, 770), "2")
    background = document.new_page()
    background.insert_text((72, 100), "3 BACKGROUND.")
    background.insert_text((72, 130), "Background text must not be retained.")
    document.set_toc(
        [
            [1, "Contents", 1],
            [1, "1 PURPOSE.", 2, {"kind": 1, "page": 1, "to": pymupdf.Point(72, 100)}],
            [
                1,
                "2 APPLICABILITY.",
                2,
                {"kind": 1, "page": 1, "to": pymupdf.Point(72, 300)},
            ],
            [
                1,
                "3 BACKGROUND.",
                4,
                {"kind": 1, "page": 3, "to": pymupdf.Point(72, 100)},
            ],
        ]
    )
    document.save(path)
    document.close()

    markdown, layout = extract_faa_sections(path)
    artifact = layout["_structured_artifact"]["value"]

    assert "The purpose begins here" in markdown
    assert "Applicability continues on the next page" in markdown
    assert "1 PURPOSE ........ 1" not in markdown
    assert "RUNNING HEADER" not in markdown
    assert "\n2\n" not in markdown
    assert "3 BACKGROUND" not in markdown
    assert "Background text must not be retained" not in markdown
    assert artifact["purpose"]["page_range"] == [2, 2]
    assert artifact["applicability"]["page_range"] == [2, 4]


def test_faa_profile_fails_when_required_headings_are_missing(tmp_path: Path) -> None:
    path = tmp_path / "FAA_missing.pdf"
    _save_pdf(path, [[((72, 72), "Contents"), ((72, 100), "1 PURPOSE .... 1")]])
    with pytest.raises(ValueError, match="PURPOSE heading"):
        extract_faa_sections(path)
